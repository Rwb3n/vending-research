#!/usr/bin/env python3
"""Generate the Snack Choice vending business research workbook."""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# ── Styles ──────────────────────────────────────────────────────────────

BLUE_FONT = Font(color="0000FF", bold=False, size=11)
BLACK_FONT = Font(color="000000", size=11)
BOLD_FONT = Font(bold=True, size=11)
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=12)

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

GBP_FORMAT = '£#,##0'
GBP_PENCE_FORMAT = '£#,##0.00'
PCT_FORMAT = '0.0%'


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dropdown(ws, validation_range, options):
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(options)}"',
        allow_blank=True,
    )
    dv.error = "Please select from the dropdown list"
    dv.errorTitle = "Invalid Entry"
    ws.add_data_validation(dv)
    dv.add(validation_range)
    return dv


# ── Data loading ────────────────────────────────────────────────────────

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")


def load_json(filename):
    path = os.path.join(DATA_DIR, filename)
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return []


# ── Sheet builders ──────────────────────────────────────────────────────


def build_product_catalog(wb):
    ws = wb.create_sheet("Product Catalog")
    headers = [
        "Category", "Product Name", "Brand", "Pack Size",
        "Wholesale Unit Cost (£)", "Wholesale Source", "Vend Price (£)",
        "Gross Margin (£)", "Gross Margin (%)", "Card Fee (£)",
        "Net Margin (£)", "Net Margin (%)", "Location Fit", "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))
    set_col_widths(ws, [16, 24, 14, 10, 22, 18, 16, 16, 16, 12, 14, 14, 14, 30])

    # Dropdowns
    add_dropdown(ws, "A2:A200", [
        "Confectionery", "Crisps", "Cold Drinks", "Water",
        "Protein/Health", "Hot Drinks", "Other",
    ])
    add_dropdown(ws, "F2:F200", [
        "Booker", "Bestway", "Costco Business", "JJ Food Service",
        "Vending Superstore", "Direct from Brand", "Other",
    ])
    add_dropdown(ws, "M2:M200", [
        "Office", "Gym", "Leisure", "Education", "Universal",
    ])

    # Load data
    products = load_json("products.json")
    for i, p in enumerate(products, 2):
        ws.cell(row=i, column=1, value=p.get("category", "")).font = BLACK_FONT
        ws.cell(row=i, column=2, value=p.get("name", "")).font = BLACK_FONT
        ws.cell(row=i, column=3, value=p.get("brand", "")).font = BLACK_FONT
        ws.cell(row=i, column=4, value=p.get("pack_size", "")).font = BLACK_FONT
        cell_e = ws.cell(row=i, column=5, value=p.get("wholesale_cost"))
        cell_e.font = BLUE_FONT
        cell_e.number_format = GBP_PENCE_FORMAT
        ws.cell(row=i, column=6, value=p.get("source", "")).font = BLACK_FONT
        cell_g = ws.cell(row=i, column=7, value=p.get("vend_price"))
        cell_g.font = BLUE_FONT
        cell_g.number_format = GBP_PENCE_FORMAT
        # Formulas
        ws.cell(row=i, column=8).value = f"=G{i}-E{i}"
        ws.cell(row=i, column=8).number_format = GBP_PENCE_FORMAT
        ws.cell(row=i, column=9).value = f"=(G{i}-E{i})/G{i}"
        ws.cell(row=i, column=9).number_format = PCT_FORMAT
        ws.cell(row=i, column=10).value = f"=G{i}*0.029"
        ws.cell(row=i, column=10).number_format = GBP_PENCE_FORMAT
        ws.cell(row=i, column=11).value = f"=H{i}-J{i}"
        ws.cell(row=i, column=11).number_format = GBP_PENCE_FORMAT
        ws.cell(row=i, column=12).value = f"=K{i}/G{i}"
        ws.cell(row=i, column=12).number_format = PCT_FORMAT
        ws.cell(row=i, column=13, value=p.get("location_fit", "")).font = BLACK_FONT
        ws.cell(row=i, column=14, value=p.get("notes", "")).font = BLACK_FONT

    # Add formulas for empty rows too (up to row 60)
    max_data = max(len(products) + 2, 2)
    for i in range(max_data, 62):
        ws.cell(row=i, column=8).value = f"=IF(G{i}=\"\",\"\",G{i}-E{i})"
        ws.cell(row=i, column=8).number_format = GBP_PENCE_FORMAT
        ws.cell(row=i, column=9).value = f"=IF(G{i}=\"\",\"\",((G{i}-E{i})/G{i}))"
        ws.cell(row=i, column=9).number_format = PCT_FORMAT
        ws.cell(row=i, column=10).value = f"=IF(G{i}=\"\",\"\",G{i}*0.029)"
        ws.cell(row=i, column=10).number_format = GBP_PENCE_FORMAT
        ws.cell(row=i, column=11).value = f"=IF(H{i}=\"\",\"\",H{i}-J{i})"
        ws.cell(row=i, column=11).number_format = GBP_PENCE_FORMAT
        ws.cell(row=i, column=12).value = f"=IF(K{i}=\"\",\"\",K{i}/G{i})"
        ws.cell(row=i, column=12).number_format = PCT_FORMAT

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{max(len(products)+1, 60)}"
    return ws


def build_machine_specs(wb):
    ws = wb.create_sheet("Machine Specs")
    headers = [
        "Machine Name", "Manufacturer", "Type", "Condition",
        "Purchase Price (£)", "VAT (£)", "Total Price (£)",
        "Lease Price (£/month)", "Capacity (selections)",
        "Payment Systems", "Cashless Ready", "Telemetry Built-in",
        "Energy Rating", "Dimensions (W×D×H cm)", "Weight (kg)",
        "Refrigerated", "Supplier", "Supplier Contact",
        "Warranty", "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))
    set_col_widths(ws, [
        20, 16, 10, 14, 18, 12, 16, 18, 18,
        18, 14, 16, 12, 22, 12, 12, 18, 24, 18, 30,
    ])

    add_dropdown(ws, "C2:C50", ["Snack", "Drink", "Combo", "Coffee", "Fresh Food", "Other"])
    add_dropdown(ws, "D2:D50", ["New", "Refurbished", "Used"])
    add_dropdown(ws, "K2:K50", ["Yes", "No", "Retrofit Needed"])
    add_dropdown(ws, "L2:L50", ["Yes", "No"])
    add_dropdown(ws, "P2:P50", ["Yes", "No"])

    machines = load_json("machines.json")
    for i, m in enumerate(machines, 2):
        ws.cell(row=i, column=1, value=m.get("name", ""))
        ws.cell(row=i, column=2, value=m.get("manufacturer", ""))
        ws.cell(row=i, column=3, value=m.get("type", ""))
        ws.cell(row=i, column=4, value=m.get("condition", ""))
        cell_e = ws.cell(row=i, column=5, value=m.get("price"))
        cell_e.font = BLUE_FONT
        cell_e.number_format = GBP_FORMAT
        ws.cell(row=i, column=6).value = f"=E{i}*0.2"
        ws.cell(row=i, column=6).number_format = GBP_FORMAT
        ws.cell(row=i, column=7).value = f"=E{i}+F{i}"
        ws.cell(row=i, column=7).number_format = GBP_FORMAT
        cell_h = ws.cell(row=i, column=8, value=m.get("lease_price"))
        cell_h.font = BLUE_FONT
        if m.get("lease_price"):
            cell_h.number_format = GBP_FORMAT
        ws.cell(row=i, column=9, value=m.get("capacity"))
        ws.cell(row=i, column=10, value=m.get("payment_systems", ""))
        ws.cell(row=i, column=11, value=m.get("cashless_ready", ""))
        ws.cell(row=i, column=12, value=m.get("telemetry", ""))
        ws.cell(row=i, column=13, value=m.get("energy_rating", ""))
        ws.cell(row=i, column=14, value=m.get("dimensions", ""))
        ws.cell(row=i, column=15, value=m.get("weight"))
        ws.cell(row=i, column=16, value=m.get("refrigerated", ""))
        ws.cell(row=i, column=17, value=m.get("supplier", ""))
        ws.cell(row=i, column=18, value=m.get("supplier_contact", ""))
        ws.cell(row=i, column=19, value=m.get("warranty", ""))
        ws.cell(row=i, column=20, value=m.get("notes", ""))

    # VAT/Total formulas for empty rows
    for i in range(len(machines) + 2, 30):
        ws.cell(row=i, column=6).value = f"=IF(E{i}=\"\",\"\",E{i}*0.2)"
        ws.cell(row=i, column=6).number_format = GBP_FORMAT
        ws.cell(row=i, column=7).value = f"=IF(E{i}=\"\",\"\",E{i}+F{i})"
        ws.cell(row=i, column=7).number_format = GBP_FORMAT

    ws.freeze_panes = "A2"
    return ws


def build_supplier_directory(wb):
    ws = wb.create_sheet("Supplier Directory")
    headers = [
        "Supplier Name", "Category", "Products/Services", "Website",
        "Phone", "Email", "Location", "Account Required",
        "Minimum Order", "Delivery", "Payment Terms", "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))
    set_col_widths(ws, [22, 18, 28, 28, 18, 28, 18, 16, 14, 20, 18, 30])

    add_dropdown(ws, "B2:B100", [
        "Wholesaler", "Machine Supplier", "Payment Systems",
        "Insurance", "Maintenance/Parts", "Trade Association",
        "Legal/Compliance", "Other",
    ])
    add_dropdown(ws, "H2:H100", ["Yes", "No", "Trade Account Only"])

    suppliers = load_json("suppliers.json")
    for i, s in enumerate(suppliers, 2):
        ws.cell(row=i, column=1, value=s.get("name", ""))
        ws.cell(row=i, column=2, value=s.get("category", ""))
        ws.cell(row=i, column=3, value=s.get("products", ""))
        ws.cell(row=i, column=4, value=s.get("website", ""))
        ws.cell(row=i, column=5, value=s.get("phone", ""))
        ws.cell(row=i, column=6, value=s.get("email", ""))
        ws.cell(row=i, column=7, value=s.get("location", ""))
        ws.cell(row=i, column=8, value=s.get("account_required", ""))
        ws.cell(row=i, column=9, value=s.get("min_order", ""))
        ws.cell(row=i, column=10, value=s.get("delivery", ""))
        ws.cell(row=i, column=11, value=s.get("payment_terms", ""))
        ws.cell(row=i, column=12, value=s.get("notes", ""))

    ws.freeze_panes = "A2"
    return ws


def build_competitor_analysis(wb):
    ws = wb.create_sheet("Competitor Analysis")
    headers = [
        "Operator Name", "Type", "HQ Location", "UK Coverage",
        "London Presence", "Machine Count (est.)", "Service Model",
        "Target Sectors", "Key Differentiator", "Pricing Model",
        "Website", "Threat Level", "Opportunity", "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))
    set_col_widths(ws, [22, 14, 16, 18, 16, 18, 22, 22, 24, 22, 24, 14, 28, 30])

    add_dropdown(ws, "B2:B50", ["National", "Regional", "Independent", "Micro-Market"])
    add_dropdown(ws, "E2:E50", ["Strong", "Moderate", "Weak", "Unknown"])
    add_dropdown(ws, "L2:L50", ["High", "Medium", "Low"])

    competitors = load_json("competitors.json")
    for i, c in enumerate(competitors, 2):
        ws.cell(row=i, column=1, value=c.get("name", ""))
        ws.cell(row=i, column=2, value=c.get("type", ""))
        ws.cell(row=i, column=3, value=c.get("hq", ""))
        ws.cell(row=i, column=4, value=c.get("uk_coverage", ""))
        ws.cell(row=i, column=5, value=c.get("london_presence", ""))
        ws.cell(row=i, column=6, value=c.get("machine_count", ""))
        ws.cell(row=i, column=7, value=c.get("service_model", ""))
        ws.cell(row=i, column=8, value=c.get("target_sectors", ""))
        ws.cell(row=i, column=9, value=c.get("differentiator", ""))
        ws.cell(row=i, column=10, value=c.get("pricing_model", ""))
        ws.cell(row=i, column=11, value=c.get("website", ""))
        ws.cell(row=i, column=12, value=c.get("threat_level", ""))
        ws.cell(row=i, column=13, value=c.get("opportunity", ""))
        ws.cell(row=i, column=14, value=c.get("notes", ""))

    ws.freeze_panes = "A2"
    return ws


def build_startup_costs(wb):
    ws = wb.create_sheet("Startup Costs")
    set_col_widths(ws, [36, 22, 16, 18, 30])

    # Title
    ws["A1"] = "STARTUP COST CALCULATOR"
    ws["A1"].font = TITLE_FONT

    # Scenario selector
    ws["A2"] = "Number of machines:"
    ws["A2"].font = BOLD_FONT
    ws["B2"] = 2
    ws["B2"].font = BLUE_FONT
    ws["B2"].fill = INPUT_FILL
    ws["C2"] = "machines"

    # Section: Capital Costs
    row = 4
    for col_letter in ["A", "B", "C", "D"]:
        ws[f"{col_letter}{row}"].fill = SECTION_FILL
    ws[f"A{row}"] = "CAPITAL COSTS"
    ws[f"A{row}"].font = SECTION_FONT
    ws[f"B{row}"] = "Per Machine (£)"
    ws[f"B{row}"].font = BOLD_FONT
    ws[f"C{row}"] = "Quantity"
    ws[f"C{row}"].font = BOLD_FONT
    ws[f"D{row}"] = "Total (£)"
    ws[f"D{row}"].font = BOLD_FONT

    capital_items = [
        ("Vending Machine (refurb)", 1750, True),
        ("VAT on Machine", None, False),  # formula
        ("Nayax VPOS Touch", 450, True),
        ("VAT on Nayax", None, False),
        ("Delivery & Installation", 150, True),
        ("Initial Stock Fill", 200, True),
    ]

    r = 5
    for label, default, is_input in capital_items:
        ws.cell(row=r, column=1, value=label).border = THIN_BORDER
        cell_b = ws.cell(row=r, column=2)
        if is_input:
            cell_b.value = default
            cell_b.font = BLUE_FONT
            cell_b.fill = INPUT_FILL
        elif label == "VAT on Machine":
            cell_b.value = "=B5*0.2"
        elif label == "VAT on Nayax":
            cell_b.value = "=B7*0.2"
        cell_b.number_format = GBP_FORMAT
        cell_b.border = THIN_BORDER

        ws.cell(row=r, column=3, value=f"=$B$2").border = THIN_BORDER
        ws.cell(row=r, column=4).value = f"=B{r}*C{r}"
        ws.cell(row=r, column=4).number_format = GBP_FORMAT
        ws.cell(row=r, column=4).border = THIN_BORDER
        r += 1

    # Subtotal Capital
    ws.cell(row=r, column=1, value="Subtotal Capital").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.cell(row=r, column=4).value = f"=SUM(D5:D{r-1})"
    ws.cell(row=r, column=4).number_format = GBP_FORMAT
    ws.cell(row=r, column=4).font = BOLD_FONT
    ws.cell(row=r, column=4).border = THIN_BORDER
    capital_subtotal_row = r
    r += 2

    # Section: Setup Costs (one-off)
    for col_letter in ["A", "B", "C", "D"]:
        ws[f"{col_letter}{r}"].fill = SECTION_FILL
    ws.cell(row=r, column=1, value="SETUP COSTS (ONE-OFF)").font = SECTION_FONT
    ws.cell(row=r, column=2, value="Cost (£)").font = BOLD_FONT
    ws.cell(row=r, column=3, value="Qty").font = BOLD_FONT
    ws.cell(row=r, column=4, value="Total (£)").font = BOLD_FONT
    r += 1

    setup_items = [
        ("Company Registration (Ltd)", 50, 1, "Companies House"),
        ("Food Business Registration", 0, 1, "Free, local council, 28 days before trading"),
        ("Food Hygiene Certificate", 150, 1, "Level 2 online course"),
        ("Public Liability Insurance (annual)", 150, 1, "£1-2M cover"),
        ("Product Liability Insurance (annual)", 100, 1, "Often bundled with above"),
        ("PAT Testing", 50, None, "Per machine"),  # qty = B2
        ("Nayax KYC/Setup Fee", 45, 1, "Per Nayax account"),
        ("Branding/Wrapping (optional)", 200, None, "Custom vinyl wrap per machine"),
    ]
    setup_start = r
    for label, cost, qty, note in setup_items:
        ws.cell(row=r, column=1, value=label).border = THIN_BORDER
        cell_b = ws.cell(row=r, column=2, value=cost)
        cell_b.font = BLUE_FONT
        cell_b.fill = INPUT_FILL
        cell_b.number_format = GBP_FORMAT
        cell_b.border = THIN_BORDER
        if qty is not None:
            ws.cell(row=r, column=3, value=qty).border = THIN_BORDER
        else:
            ws.cell(row=r, column=3, value="=$B$2").border = THIN_BORDER
        ws.cell(row=r, column=4).value = f"=B{r}*C{r}"
        ws.cell(row=r, column=4).number_format = GBP_FORMAT
        ws.cell(row=r, column=4).border = THIN_BORDER
        ws.cell(row=r, column=5, value=note).font = Font(italic=True, color="808080", size=10)
        r += 1

    ws.cell(row=r, column=1, value="Subtotal Setup").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.cell(row=r, column=4).value = f"=SUM(D{setup_start}:D{r-1})"
    ws.cell(row=r, column=4).number_format = GBP_FORMAT
    ws.cell(row=r, column=4).font = BOLD_FONT
    ws.cell(row=r, column=4).border = THIN_BORDER
    setup_subtotal_row = r
    r += 2

    # Section: Monthly Operating Costs
    for col_letter in ["A", "B", "C", "D"]:
        ws[f"{col_letter}{r}"].fill = SECTION_FILL
    ws.cell(row=r, column=1, value="MONTHLY OPERATING COSTS").font = SECTION_FONT
    ws.cell(row=r, column=2, value="Per Machine/Month (£)").font = BOLD_FONT
    ws.cell(row=r, column=4, value="Total/Month (£)").font = BOLD_FONT
    r += 1

    opex_items = [
        ("Nayax Monthly Fee", 10, False),
        ("Site Rent / Commission", 100, True),
        ("Stock Replenishment", 400, True),
        ("Travel/Fuel", 80, True),
        ("Electricity (if not site-covered)", 0, True),
        ("Maintenance/Repairs Fund", 30, True),
        ("Misc/Phone/Admin", 20, False),
    ]
    opex_start = r
    for label, default, is_input in opex_items:
        ws.cell(row=r, column=1, value=label).border = THIN_BORDER
        cell_b = ws.cell(row=r, column=2, value=default)
        cell_b.number_format = GBP_FORMAT
        cell_b.border = THIN_BORDER
        if is_input:
            cell_b.font = BLUE_FONT
            cell_b.fill = INPUT_FILL
        if label == "Misc/Phone/Admin":
            ws.cell(row=r, column=4).value = f"=B{r}"
        else:
            ws.cell(row=r, column=4).value = f"=B{r}*$B$2"
        ws.cell(row=r, column=4).number_format = GBP_FORMAT
        ws.cell(row=r, column=4).border = THIN_BORDER
        r += 1

    ws.cell(row=r, column=1, value="Total Monthly Opex").font = BOLD_FONT
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.cell(row=r, column=4).value = f"=SUM(D{opex_start}:D{r-1})"
    ws.cell(row=r, column=4).number_format = GBP_FORMAT
    ws.cell(row=r, column=4).font = BOLD_FONT
    ws.cell(row=r, column=4).border = THIN_BORDER
    opex_total_row = r
    r += 2

    # Section: Summary
    for col_letter in ["A", "B", "C", "D"]:
        ws[f"{col_letter}{r}"].fill = SECTION_FILL
    ws.cell(row=r, column=1, value="SUMMARY").font = SECTION_FONT
    r += 1

    summary_items = [
        ("Total Upfront Investment", f"=D{capital_subtotal_row}+D{setup_subtotal_row}", False),
        ("Monthly Operating Cost", f"=D{opex_total_row}", False),
        ("Est. Monthly Revenue per Machine", 2000, True),
        ("Est. Monthly Revenue Total", None, False),  # special
        ("Est. Monthly Profit", None, False),  # special
        ("Months to Breakeven", None, False),  # special
    ]
    summary_start = r
    for label, val, is_input in summary_items:
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        cell_d = ws.cell(row=r, column=4)
        cell_d.border = THIN_BORDER
        if label == "Est. Monthly Revenue per Machine":
            cell_b = ws.cell(row=r, column=2, value=val)
            cell_b.font = BLUE_FONT
            cell_b.fill = INPUT_FILL
            cell_b.number_format = GBP_FORMAT
            cell_b.border = THIN_BORDER
            cell_d.value = f"=B{r}*$B$2"
        elif label == "Est. Monthly Revenue Total":
            cell_d.value = f"=D{r-1}"
        elif label == "Est. Monthly Profit":
            cell_d.value = f"=D{r-1}-D{summary_start+1}"
        elif label == "Months to Breakeven":
            cell_d.value = f"=D{summary_start}/D{r-1}"
            cell_d.number_format = "0.0"
            r += 1
            continue
        else:
            cell_d.value = val
        cell_d.number_format = GBP_FORMAT
        cell_d.font = BOLD_FONT
        r += 1

    ws.freeze_panes = "A3"
    return ws


def build_location_tracker(wb):
    ws = wb.create_sheet("Location Tracker")
    headers = [
        "Site Name", "Address", "Postcode", "Borough", "Site Type",
        "Contact Name", "Contact Phone", "Contact Email",
        "First Contact Date", "Status", "Est. Daily Footfall",
        "Existing Vending", "Current Provider", "Contract Expiry",
        "Space Available", "Power Supply Nearby",
        "Proposed Deal Structure", "Est. Monthly Rent (£)",
        "Revenue Share (%)", "Est. Weekly Revenue (£)",
        "Score (1-10)", "Notes",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))
    set_col_widths(ws, [
        22, 30, 12, 16, 18, 18, 16, 24, 16, 16, 16,
        14, 18, 14, 14, 16, 22, 18, 14, 20, 12, 30,
    ])

    add_dropdown(ws, "E2:E200", [
        "Office", "Gym", "Leisure Centre", "Education",
        "Hospital/Healthcare", "Retail", "Residential",
        "Industrial/Warehouse", "Transport Hub", "Other",
    ])
    add_dropdown(ws, "J2:J200", [
        "Lead", "Contacted", "Meeting Booked", "Negotiating",
        "Agreed", "Rejected", "On Hold",
    ])
    add_dropdown(ws, "L2:L200", ["Yes", "No", "Unknown"])
    add_dropdown(ws, "O2:O200", ["Yes", "No", "TBC"])
    add_dropdown(ws, "P2:P200", ["Yes", "No", "TBC"])

    # Conditional formatting on Status column (J)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    ws.conditional_formatting.add("J2:J200",
        CellIsRule(operator="equal", formula=['"Agreed"'], fill=green_fill))
    ws.conditional_formatting.add("J2:J200",
        CellIsRule(operator="equal", formula=['"Rejected"'], fill=red_fill))
    ws.conditional_formatting.add("J2:J200",
        CellIsRule(operator="equal", formula=['"Negotiating"'], fill=yellow_fill))
    ws.conditional_formatting.add("J2:J200",
        CellIsRule(operator="equal", formula=['"Lead"'], fill=blue_fill))

    ws.freeze_panes = "A2"
    return ws


def build_weekly_pnl(wb):
    ws = wb.create_sheet("Weekly P&L")
    set_col_widths(ws, [16, 18, 18, 18, 16, 16, 14, 18])

    # Header section
    ws["A1"] = "Machine ID:"
    ws["A1"].font = BOLD_FONT
    ws["B1"].font = BLUE_FONT
    ws["B1"].fill = INPUT_FILL
    ws["A2"] = "Location:"
    ws["A2"].font = BOLD_FONT
    ws["B2"].font = BLUE_FONT
    ws["B2"].fill = INPUT_FILL
    ws["A3"] = "Week Commencing:"
    ws["A3"].font = BOLD_FONT
    ws["B3"].font = BLUE_FONT
    ws["B3"].fill = INPUT_FILL
    ws["B3"].number_format = "DD/MM/YYYY"

    # Revenue headers
    r = 5
    rev_headers = [
        "Day", "Cash Revenue (£)", "Card Revenue (£)", "Total Revenue (£)",
        "Units Sold (Cash)", "Units Sold (Card)", "Total Units", "Avg Vend Price (£)",
    ]
    for col, h in enumerate(rev_headers, 1):
        ws.cell(row=r, column=col, value=h)
    style_header_row(ws, r, len(rev_headers))

    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    for i, day in enumerate(days):
        row = r + 1 + i
        ws.cell(row=row, column=1, value=day)
        ws.cell(row=row, column=2).number_format = GBP_PENCE_FORMAT
        ws.cell(row=row, column=2).font = BLUE_FONT
        ws.cell(row=row, column=2).fill = INPUT_FILL
        ws.cell(row=row, column=3).number_format = GBP_PENCE_FORMAT
        ws.cell(row=row, column=3).font = BLUE_FONT
        ws.cell(row=row, column=3).fill = INPUT_FILL
        ws.cell(row=row, column=4).value = f"=B{row}+C{row}"
        ws.cell(row=row, column=4).number_format = GBP_PENCE_FORMAT
        ws.cell(row=row, column=5).font = BLUE_FONT
        ws.cell(row=row, column=5).fill = INPUT_FILL
        ws.cell(row=row, column=6).font = BLUE_FONT
        ws.cell(row=row, column=6).fill = INPUT_FILL
        ws.cell(row=row, column=7).value = f"=E{row}+F{row}"
        ws.cell(row=row, column=8).value = f"=IF(G{row}=0,\"\",D{row}/G{row})"
        ws.cell(row=row, column=8).number_format = GBP_PENCE_FORMAT

    # Week total
    total_row = r + 8
    ws.cell(row=total_row, column=1, value="WEEK TOTAL").font = BOLD_FONT
    for col in [2, 3, 4]:
        ws.cell(row=total_row, column=col).value = f"=SUM({get_column_letter(col)}{r+1}:{get_column_letter(col)}{total_row-1})"
        ws.cell(row=total_row, column=col).number_format = GBP_PENCE_FORMAT
        ws.cell(row=total_row, column=col).font = BOLD_FONT
    for col in [5, 6, 7]:
        ws.cell(row=total_row, column=col).value = f"=SUM({get_column_letter(col)}{r+1}:{get_column_letter(col)}{total_row-1})"
        ws.cell(row=total_row, column=col).font = BOLD_FONT
    ws.cell(row=total_row, column=8).value = f"=IF(G{total_row}=0,\"\",D{total_row}/G{total_row})"
    ws.cell(row=total_row, column=8).number_format = GBP_PENCE_FORMAT
    ws.cell(row=total_row, column=8).font = BOLD_FONT

    # Cost section
    cr = total_row + 2
    for col_letter in ["A", "B"]:
        ws[f"{col_letter}{cr}"].fill = SECTION_FILL
    ws.cell(row=cr, column=1, value="COSTS").font = SECTION_FONT
    ws.cell(row=cr, column=2, value="Amount (£)").font = BOLD_FONT
    cr += 1

    cost_items = [
        ("Stock Purchased", None, True),
        ("Card Processing Fees", f"=SUM(C{r+1}:C{total_row-1})*0.029", False),
        ("Nayax Monthly (pro-rata weekly)", "=10/4.33", False),
        ("Site Rent (pro-rata weekly)", None, True),
        ("Travel/Fuel", None, True),
        ("Other", None, True),
    ]
    cost_start = cr
    for label, formula, is_input in cost_items:
        ws.cell(row=cr, column=1, value=label).border = THIN_BORDER
        cell_b = ws.cell(row=cr, column=2)
        cell_b.border = THIN_BORDER
        cell_b.number_format = GBP_PENCE_FORMAT
        if is_input:
            cell_b.font = BLUE_FONT
            cell_b.fill = INPUT_FILL
        else:
            cell_b.value = formula
        cr += 1

    ws.cell(row=cr, column=1, value="Total Costs").font = BOLD_FONT
    ws.cell(row=cr, column=1).border = THIN_BORDER
    ws.cell(row=cr, column=2).value = f"=SUM(B{cost_start}:B{cr-1})"
    ws.cell(row=cr, column=2).number_format = GBP_PENCE_FORMAT
    ws.cell(row=cr, column=2).font = BOLD_FONT
    ws.cell(row=cr, column=2).border = THIN_BORDER
    total_costs_row = cr
    cr += 2

    # Summary section
    for col_letter in ["A", "B"]:
        ws[f"{col_letter}{cr}"].fill = SECTION_FILL
    ws.cell(row=cr, column=1, value="WEEKLY SUMMARY").font = SECTION_FONT
    cr += 1

    ws.cell(row=cr, column=1, value="Gross Revenue").font = BOLD_FONT
    ws.cell(row=cr, column=2).value = f"=D{total_row}"
    ws.cell(row=cr, column=2).number_format = GBP_PENCE_FORMAT
    ws.cell(row=cr, column=2).font = BOLD_FONT
    cr += 1

    ws.cell(row=cr, column=1, value="Total Costs").font = BOLD_FONT
    ws.cell(row=cr, column=2).value = f"=B{total_costs_row}"
    ws.cell(row=cr, column=2).number_format = GBP_PENCE_FORMAT
    ws.cell(row=cr, column=2).font = BOLD_FONT
    cr += 1

    ws.cell(row=cr, column=1, value="Net Profit").font = BOLD_FONT
    ws.cell(row=cr, column=2).value = f"=B{cr-2}-B{cr-1}"
    ws.cell(row=cr, column=2).number_format = GBP_PENCE_FORMAT
    ws.cell(row=cr, column=2).font = BOLD_FONT
    cr += 1

    ws.cell(row=cr, column=1, value="Gross Margin %").font = BOLD_FONT
    ws.cell(row=cr, column=2).value = f"=IF(B{cr-3}=0,\"\",(B{cr-3}-B{cost_start})/B{cr-3})"
    ws.cell(row=cr, column=2).number_format = PCT_FORMAT
    ws.cell(row=cr, column=2).font = BOLD_FONT

    return ws


def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0)  # Insert at position 0
    set_col_widths(ws, [30, 20, 20, 20, 20])

    ws["A1"] = "THE SNACK CHOICE — LONDON EXPANSION"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = "Business Intelligence Dashboard"
    ws["A2"].font = Font(italic=True, size=12, color="808080")

    # Key metrics section
    r = 4
    ws.cell(row=r, column=1, value="KEY METRICS").font = SECTION_FONT
    for col_letter in ["A", "B", "C"]:
        ws[f"{col_letter}{r}"].fill = SECTION_FILL
    r += 1

    metrics = [
        ("Number of Machines", "='Startup Costs'!B2"),
        ("Total Upfront Investment", "='Startup Costs'!D{cap_setup}"),
        ("Monthly Operating Cost", "='Startup Costs'!D{opex}"),
        ("Est. Monthly Revenue", "='Startup Costs'!D{rev}"),
        ("Est. Monthly Profit", "='Startup Costs'!D{profit}"),
        ("Months to Breakeven", "='Startup Costs'!D{breakeven}"),
    ]

    # We need to reference the actual rows from startup costs
    # These are approximate — the formulas will reference the right cells
    # We'll use named references in the summary section
    dashboard_metrics = [
        ("Number of Machines", "='Startup Costs'!B2", "0"),
        ("Total Upfront Investment", "='Startup Costs'!D35", GBP_FORMAT),
        ("Monthly Operating Cost", "='Startup Costs'!D32", GBP_FORMAT),
        ("Est. Monthly Revenue", "='Startup Costs'!D38", GBP_FORMAT),
        ("Est. Monthly Profit", "='Startup Costs'!D39", GBP_FORMAT),
        ("Months to Breakeven", "='Startup Costs'!D40", "0.0"),
    ]

    for label, formula, fmt in dashboard_metrics:
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        cell = ws.cell(row=r, column=2)
        cell.border = THIN_BORDER
        cell.number_format = fmt
        cell.value = formula
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="PRODUCT CATALOG SUMMARY").font = SECTION_FONT
    for col_letter in ["A", "B", "C"]:
        ws[f"{col_letter}{r}"].fill = SECTION_FILL
    r += 1

    ws.cell(row=r, column=1, value="Total Products Listed")
    ws.cell(row=r, column=2).value = "=COUNTA('Product Catalog'!B2:B200)"
    r += 1
    ws.cell(row=r, column=1, value="Avg Gross Margin %")
    ws.cell(row=r, column=2).value = "=AVERAGE('Product Catalog'!I2:I200)"
    ws.cell(row=r, column=2).number_format = PCT_FORMAT
    r += 1
    ws.cell(row=r, column=1, value="Avg Net Margin %")
    ws.cell(row=r, column=2).value = "=AVERAGE('Product Catalog'!L2:L200)"
    ws.cell(row=r, column=2).number_format = PCT_FORMAT
    r += 1
    ws.cell(row=r, column=1, value="Min Net Margin %")
    ws.cell(row=r, column=2).value = "=MIN('Product Catalog'!L2:L200)"
    ws.cell(row=r, column=2).number_format = PCT_FORMAT
    r += 1
    ws.cell(row=r, column=1, value="Max Net Margin %")
    ws.cell(row=r, column=2).value = "=MAX('Product Catalog'!L2:L200)"
    ws.cell(row=r, column=2).number_format = PCT_FORMAT

    r += 2
    ws.cell(row=r, column=1, value="LOCATION PIPELINE").font = SECTION_FONT
    for col_letter in ["A", "B", "C"]:
        ws[f"{col_letter}{r}"].fill = SECTION_FILL
    r += 1

    statuses = ["Lead", "Contacted", "Meeting Booked", "Negotiating", "Agreed", "Rejected"]
    for status in statuses:
        ws.cell(row=r, column=1, value=status)
        ws.cell(row=r, column=2).value = f'=COUNTIF(\'Location Tracker\'!J2:J200,"{status}")'
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="MACHINES RESEARCHED").font = SECTION_FONT
    for col_letter in ["A", "B", "C"]:
        ws[f"{col_letter}{r}"].fill = SECTION_FILL
    r += 1
    ws.cell(row=r, column=1, value="Total Machines Listed")
    ws.cell(row=r, column=2).value = "=COUNTA('Machine Specs'!A2:A50)"
    r += 1
    ws.cell(row=r, column=1, value="Avg Purchase Price (ex VAT)")
    ws.cell(row=r, column=2).value = "=AVERAGE('Machine Specs'!E2:E50)"
    ws.cell(row=r, column=2).number_format = GBP_FORMAT
    r += 1
    ws.cell(row=r, column=1, value="Min Purchase Price (ex VAT)")
    ws.cell(row=r, column=2).value = "=MIN('Machine Specs'!E2:E50)"
    ws.cell(row=r, column=2).number_format = GBP_FORMAT
    r += 1
    ws.cell(row=r, column=1, value="Max Purchase Price (ex VAT)")
    ws.cell(row=r, column=2).value = "=MAX('Machine Specs'!E2:E50)"
    ws.cell(row=r, column=2).number_format = GBP_FORMAT

    r += 2
    ws.cell(row=r, column=1, value="COMPETITORS").font = SECTION_FONT
    for col_letter in ["A", "B", "C"]:
        ws[f"{col_letter}{r}"].fill = SECTION_FILL
    r += 1
    ws.cell(row=r, column=1, value="Total Competitors Tracked")
    ws.cell(row=r, column=2).value = "=COUNTA('Competitor Analysis'!A2:A50)"
    r += 1
    ws.cell(row=r, column=1, value="High Threat")
    ws.cell(row=r, column=2).value = '=COUNTIF(\'Competitor Analysis\'!L2:L50,"High")'
    r += 1
    ws.cell(row=r, column=1, value="Medium Threat")
    ws.cell(row=r, column=2).value = '=COUNTIF(\'Competitor Analysis\'!L2:L50,"Medium")'
    r += 1
    ws.cell(row=r, column=1, value="Low Threat")
    ws.cell(row=r, column=2).value = '=COUNTIF(\'Competitor Analysis\'!L2:L50,"Low")'

    ws.freeze_panes = "A4"
    return ws


def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Build all sheets
    build_dashboard(wb)
    build_product_catalog(wb)
    build_machine_specs(wb)
    build_supplier_directory(wb)
    build_competitor_analysis(wb)
    build_startup_costs(wb)
    build_location_tracker(wb)
    build_weekly_pnl(wb)

    output = os.path.join(os.path.dirname(__file__), "TheSnackChoice_London_Research.xlsx")
    wb.save(output)
    print(f"Generated: {output}")

    # Print summary
    for ws in wb.worksheets:
        print(f"  {ws.title}: {ws.max_row} rows × {ws.max_column} cols")


if __name__ == "__main__":
    main()
