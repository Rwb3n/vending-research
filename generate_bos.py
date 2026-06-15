#!/usr/bin/env python3
"""
---
id: bos.generate_workbook
purpose: Build the Business Operating System workbook (Excel) from bos/config + bos/state.
         KPIs are computed by importing bos/tools (single source of math). Mirrors the
         market-research generate.py but for the operating layer.
inputs: bos/config/*, bos/state/*, bos/tools/kpi.py
outputs: TheSnackChoice_OperatingSystem.xlsx
usage: "python generate_bos.py"
updated: 2026-06-15
---
"""
import glob
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(HERE, "bos", "tools"))
import common  # noqa: E402
import kpi as kpi_tool  # noqa: E402

OUTPUT = os.path.join(HERE, "TheSnackChoice_OperatingSystem.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=15, color="1F3864")
SUB_FONT = Font(italic=True, color="555555")
GREEN = PatternFill("solid", fgColor="C6EFCE")
AMBER = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")
STATUS_FILL = {"green": GREEN, "amber": AMBER, "red": RED, "na": None}


def header(ws, row, cols):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title(ws, text, sub=None):
    ws["A1"] = text
    ws["A1"].font = TITLE_FONT
    if sub:
        ws["A2"] = sub
        ws["A2"].font = SUB_FONT
    return 4


def front_matter(path):
    """Parse YAML front-matter from a markdown file."""
    text = open(path).read()
    if text.startswith("---"):
        end = text.find("---", 3)
        if end != -1:
            import yaml
            return yaml.safe_load(text[3:end]) or {}
    return {}


def sheet_dashboard(wb):
    ws = wb.create_sheet("OS Dashboard")
    r = title(ws, "The Snack Choice — Operating System", "North star: revenue per operator-hour. Auto-computed; do not hand-edit.")
    kpis, weeks, machines = kpi_tool.load_inputs()
    latest = weeks[-1]
    ws.cell(row=r, column=1, value=f"Latest week: {latest['week']}").font = Font(bold=True)
    r += 1
    header(ws, r, ["KPI", "Value", "Unit", "Target", "Status"])
    r += 1
    for k in kpi_tool.compute_week(kpis, latest, machines):
        ws.cell(row=r, column=1, value=k["label"] + (" ★" if k["north_star"] else ""))
        ws.cell(row=r, column=2, value=k["value"])
        ws.cell(row=r, column=3, value=k["unit"])
        ws.cell(row=r, column=4, value=k["target"])
        s = ws.cell(row=r, column=5, value=k["status"].upper())
        if STATUS_FILL.get(k["status"]):
            s.fill = STATUS_FILL[k["status"]]
        r += 1
    widths(ws, [34, 12, 12, 10, 12])


def sheet_scorecard(wb):
    ws = wb.create_sheet("Scorecard")
    r = title(ws, "Weekly Scorecard", "Raw inputs (state/scorecard.jsonl) + derived KPIs. Append via tools/scorecard.py.")
    kpis, weeks, machines = kpi_tool.load_inputs()
    kpi_labels = [k["label"] for k in kpis if k.get("source") != "machines_state"]
    header(ws, r, ["Week", "Op hrs", "Revenue £", "COGS £", "Vends", "Mach-days"] + kpi_labels)
    r += 1
    for w in weeks:
        row = [w["week"], w["operator_hours"], w["gross_revenue"], w["cogs"], w["vends"], w["machine_days"]]
        computed = {k["id"]: k for k in kpi_tool.compute_week(kpis, w, machines)}
        base = len(row)
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
        for j, k in enumerate([k for k in kpis if k.get("source") != "machines_state"]):
            cell = ws.cell(row=r, column=base + 1 + j, value=computed[k["id"]]["value"])
            st = computed[k["id"]]["status"]
            if STATUS_FILL.get(st):
                cell.fill = STATUS_FILL[st]
        r += 1
    widths(ws, [11, 8, 11, 9, 8, 11] + [16] * len(kpi_labels))


def sheet_cadence(wb):
    ws = wb.create_sheet("Cadence")
    r = title(ws, "Operating Cadence", "The rhythm. Workflows bind to these triggers.")
    data = common.load_yaml(os.path.join(common.CONFIG_DIR, "cadence.yaml"))["cadence"]
    header(ws, r, ["Cadence", "Trigger", "Purpose", "Routines"])
    r += 1
    for c in data:
        ws.cell(row=r, column=1, value=c["id"]).font = Font(bold=True)
        ws.cell(row=r, column=2, value=c["trigger"])
        ws.cell(row=r, column=3, value=c["purpose"]).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=4, value="\n".join("• " + x for x in c["routines"])).alignment = Alignment(wrap_text=True)
        r += 1
    widths(ws, [12, 22, 40, 60])


def sheet_work(wb):
    ws = wb.create_sheet("Workflows & Pipelines")
    r = title(ws, "Tasks → Pipelines → Workflows", "No roles — only work. Any task runs by agent or human.")
    wf = common.load_yaml(os.path.join(common.CONFIG_DIR, "workflows.yaml"))["workflows"]
    pl = {p["id"]: p for p in common.load_yaml(os.path.join(common.CONFIG_DIR, "pipelines.yaml"))["pipelines"]}
    header(ws, r, ["Workflow", "Cadence", "Pipelines", "Steps (tasks)", "Done when"])
    r += 1
    for w in wf:
        steps = []
        for pid in w.get("runs", []):
            steps.extend(pl.get(pid, {}).get("steps", []))
        ws.cell(row=r, column=1, value=w["id"]).font = Font(bold=True)
        ws.cell(row=r, column=2, value=w["cadence"])
        ws.cell(row=r, column=3, value=", ".join(w.get("runs", [])) or "(event-driven)")
        ws.cell(row=r, column=4, value=" → ".join(steps) or ", ".join(w.get("tasks", []))).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=5, value=w["done_when"]).alignment = Alignment(wrap_text=True)
        r += 1
    widths(ws, [20, 12, 26, 44, 50])


def sheet_playbooks(wb):
    ws = wb.create_sheet("Playbooks")
    r = title(ws, "Playbooks (SOPs)", "Human-runnable procedures. Full text in bos/config/playbooks/.")
    header(ws, r, ["Playbook", "Title", "Trigger", "Protects KPI", "File"])
    r += 1
    for path in sorted(glob.glob(os.path.join(common.CONFIG_DIR, "playbooks", "*.md"))):
        fm = front_matter(path)
        ws.cell(row=r, column=1, value=fm.get("id", "").replace("playbook.", ""))
        ws.cell(row=r, column=2, value=fm.get("title"))
        ws.cell(row=r, column=3, value=str(fm.get("trigger", ""))).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=4, value=fm.get("protects_kpi", ""))
        ws.cell(row=r, column=5, value="config/playbooks/" + os.path.basename(path))
        r += 1
    widths(ws, [22, 22, 40, 24, 34])


def sheet_policies(wb):
    ws = wb.create_sheet("Decision Policies")
    r = title(ws, "Decision Policies", "Deterministic rules. Evaluate via tools/policy.py; verdicts logged to decisions.jsonl.")
    pol = common.load_yaml(os.path.join(common.CONFIG_DIR, "policies.yaml"))["policies"]
    header(ws, r, ["Policy", "Question", "Inputs", "Rules (first match wins)", "Evidence"])
    r += 1
    for p in pol:
        rules = [f"if {ru['when']} → {ru['verdict']}" for ru in p.get("rules", [])]
        rules.append(f"else → {p.get('default', {}).get('verdict', '?')}")
        ws.cell(row=r, column=1, value=p["id"]).font = Font(bold=True)
        ws.cell(row=r, column=2, value=p["label"]).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=3, value=", ".join(p["inputs"]))
        ws.cell(row=r, column=4, value="\n".join(rules)).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=5, value=p.get("evidence", "")).alignment = Alignment(wrap_text=True)
        r += 1
    widths(ws, [20, 32, 24, 50, 40])


def sheet_machines(wb):
    ws = wb.create_sheet("Machines")
    r = title(ws, "Fleet", "Current state (bos/state/machines_state.json).")
    machines = common.load_json(os.path.join(common.STATE_DIR, "machines_state.json"))["machines"]
    header(ws, r, ["ID", "Site", "Postcode", "Status", "Fill %", "Last restock", "Planogram"])
    r += 1
    for m in machines:
        ws.cell(row=r, column=1, value=m["machine_id"])
        ws.cell(row=r, column=2, value=m["site"])
        ws.cell(row=r, column=3, value=m.get("postcode"))
        ws.cell(row=r, column=4, value=m["status"])
        fill = ws.cell(row=r, column=5, value=round(m["current_fill"] * 100))
        if m["status"] == "live" and m["current_fill"] < 0.30:
            fill.fill = RED
        ws.cell(row=r, column=6, value=m.get("last_restock"))
        ws.cell(row=r, column=7, value=m.get("planogram"))
        r += 1
    widths(ws, [8, 40, 12, 16, 10, 14, 20])


def sheet_risks(wb):
    ws = wb.create_sheet("Risk Register")
    r = title(ws, "Risk Register", "Failure modes, early-warning signals, mitigations.")
    risks = common.load_yaml(os.path.join(common.CONFIG_DIR, "risks.yaml"))["risks"]
    header(ws, r, ["ID", "Risk", "Likelihood", "Impact", "Signal", "Mitigation"])
    r += 1
    for x in risks:
        ws.cell(row=r, column=1, value=x["id"])
        ws.cell(row=r, column=2, value=x["risk"]).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=3, value=x["likelihood"])
        ws.cell(row=r, column=4, value=x["impact"])
        ws.cell(row=r, column=5, value=x["signal"]).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=6, value=x["mitigation"]).alignment = Alignment(wrap_text=True)
        r += 1
    widths(ws, [6, 34, 12, 10, 38, 50])


def sheet_planograms(wb):
    ws = wb.create_sheet("Planograms")
    r = title(ws, "Planograms", "Slot → SKU → par plan per machine. reorder.py refills to par. Par levels: spike 007.")
    planograms = common.load_yaml(os.path.join(common.CONFIG_DIR, "planograms.yaml"))["planograms"]
    header(ws, r, ["Planogram", "SKU", "Category", "Facings", "Par/facing", "Units"])
    r += 1
    for pid, pg in planograms.items():
        total = sum(ln["facings"] * ln["par_each"] for ln in pg["lines"])
        facings = sum(ln["facings"] for ln in pg["lines"])
        ws.cell(row=r, column=1, value=f"{pid} ({facings} sel, {total} items)").font = Font(bold=True)
        r += 1
        for ln in pg["lines"]:
            ws.cell(row=r, column=2, value=ln["sku"])
            ws.cell(row=r, column=3, value=ln["category"])
            ws.cell(row=r, column=4, value=ln["facings"])
            ws.cell(row=r, column=5, value=ln["par_each"])
            ws.cell(row=r, column=6, value=ln["facings"] * ln["par_each"])
            r += 1
    widths(ws, [16, 44, 16, 9, 11, 8])


def sheet_assumptions(wb):
    ws = wb.create_sheet("Assumptions & Spikes")
    r = title(ws, "Assumptions & Spikes", "Evidence or spike. SPIKE rows have an experiment in scratch/spikes/.")
    header(ws, r, ["Spike", "Title", "Status", "Unblocks"])
    r += 1
    for path in sorted(glob.glob(os.path.join(HERE, "scratch", "spikes", "*.md"))):
        fm = front_matter(path)
        ws.cell(row=r, column=1, value=fm.get("id", ""))
        ws.cell(row=r, column=2, value=fm.get("title", ""))
        ws.cell(row=r, column=3, value=fm.get("status", ""))
        ws.cell(row=r, column=4, value=", ".join(fm.get("unblocks", []) if isinstance(fm.get("unblocks"), list) else [str(fm.get("unblocks", ""))])).alignment = Alignment(wrap_text=True)
        r += 1
    widths(ws, [12, 44, 10, 50])


def main():
    wb = Workbook()
    wb.remove(wb.active)
    sheet_dashboard(wb)
    sheet_scorecard(wb)
    sheet_cadence(wb)
    sheet_work(wb)
    sheet_playbooks(wb)
    sheet_policies(wb)
    sheet_machines(wb)
    sheet_planograms(wb)
    sheet_risks(wb)
    sheet_assumptions(wb)
    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT} ({len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)})")


if __name__ == "__main__":
    main()
